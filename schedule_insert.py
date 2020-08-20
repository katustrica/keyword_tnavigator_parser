import argparse
import re
from collections import namedtuple
from typing import List
from datetime import datetime
from pathlib import Path

import PySimpleGUIQt as sg
from openpyxl import load_workbook

WFRACP = 'WFRACP'
COMPDAT = 'COMPDAT'
Keyword = namedtuple('Keyword', ['type', 'date', 'value'])

def load_excel_with_data(fractures_path: Path):
    wb = load_workbook(fractures_path, data_only=True)
    sheet = wb.active
    excel_data: List[Keyword] = []

    for row in range(1, sheet.max_row):
        date_time_value = sheet.cell(row, 2).value
        if isinstance(date_time_value, datetime):
            data_list = []
            for column in range(3, sheet.max_column + 1):
                data_list.append(str(sheet.cell(row, column).value))
            keyword_cell_value = sheet.cell(row - 1, 3).value
            if keyword_cell_value == WFRACP:
                keyword = WFRACP
            elif keyword_cell_value == COMPDAT:
                keyword = COMPDAT
            else:
                raise ValueError(f"Unknow keywoqrd: {keyword_cell_value} row - {row-1}")
            data = Keyword(keyword, date_time_value.strftime("%e %b %Y").upper(), ' '.join(data_list))
            excel_data.append(data)
    return excel_data


def paste_wfracp_in_txt(schedule_path, excel_data):
    with open(schedule_path, "r+") as f:
        lines_list = [x.rstrip() for x in f]
        for item in excel_data:
            for line_index in range(0, len(lines_list)):
                if lines_list[line_index].find(re.sub(' ', r'\t', item.date)) != -1:
                    for line in range(line_index, len(lines_list)):
                        if lines_list[line].find("DATES") != -1:
                            lines_list.insert(line, f"{item.type}\n{item.value}\n/\n")
                            break

        return lines_list


def write_file(string_list: list, filepath: Path):
    with open(filepath, 'w+') as filehandle:
        for listitem in string_list:
            filehandle.write(f'{listitem}\n')

if __name__ == "__main__":
    layout = [
        [sg.Text('Excel:', size=(10, 1)), sg.InputText(key='--EXCEL--', size=(20, 1)), sg.FileBrowse()],
        [sg.Text('SCHEDULE:', size=(10, 1)), sg.InputText(key='--TXT--', size=(20,1)), sg.FileBrowse()],
        [sg.Text('Result:', size=(10, 1)), sg.InputText(key='--RESULT--', size=(20,1)), sg.FileSaveAs()],
        [sg.Submit(), sg.Cancel()]
    ]
    window = sg.Window('Parser 2000', layout)
    while True:
        event, values = window.read()
        # print(event, values) #debug
        if event in (None, 'Exit', 'Cancel'):
            break
        if event == 'Submit':
            if values['--EXCEL--'] and values['--TXT--']:
                excelPath = Path(values['--EXCEL--'])
                txtPath = Path(values['--TXT--'])
                resultPath = None
                if values['--RESULT--']:
                    resultPath = Path(values['--RESULT--'])
                try:
                    output_path = resultPath or txtPath.parent / 'result.txt'
                    excel_data = load_excel_with_data(excelPath)
                    results_lines_list = paste_wfracp_in_txt(txtPath, excel_data)
                    write_file(results_lines_list, output_path)
                    sg.PopupOK('Done👌.')
                except Exception as e:
                    sg.PopupNonBlocking(e)
            else:
                sg.PopupNonBlocking('Please choose files.')
